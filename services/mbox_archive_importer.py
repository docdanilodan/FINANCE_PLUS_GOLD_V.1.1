from __future__ import annotations

import csv
import hashlib
import html
import json
import mailbox
import mimetypes
import os
import re
import shutil
import tempfile
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from difflib import SequenceMatcher
from email.header import decode_header, make_header
from email.message import Message
from email.utils import getaddresses, parsedate_to_datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Callable, Iterable, Sequence

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

MEDIA_EXTENSIONS = {
    ".3gp", ".avi", ".bmp", ".gif", ".heic", ".heif", ".jpeg", ".jpg",
    ".m4v", ".mkv", ".mov", ".mp4", ".mpeg", ".mpg", ".png", ".tif",
    ".tiff", ".webm", ".webp",
}
COMPANY_SUFFIX_RE = re.compile(
    r"\b([A-ZÀ-ÖØ-Ý0-9][A-ZÀ-ÖØ-Ý0-9&'’._\-/ ]{1,90}?\s+"
    r"(?:S\.?\s*R\.?\s*L\.?\s*S?\.?|S\.?\s*P\.?\s*A\.?|S\.?\s*N\.?\s*C\.?|"
    r"S\.?\s*A\.?\s*S\.?|SOCIETA\s+COOPERATIVA|COOPERATIVA))\b",
    re.IGNORECASE,
)
SAFE_COMPONENT_RE = re.compile(r"[^A-Za-z0-9À-ÖØ-öø-ÿ._()\- ]+")


class _HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.parts.append(data.strip())

    def text(self) -> str:
        return "\n".join(self.parts)


@dataclass
class ImportResult:
    source: str
    output_dir: str
    email_count: int
    attachment_count: int
    catalogued_count: int
    unclassified_count: int
    excluded_media_count: int
    duplicate_count: int
    email_csv: str
    attachment_csv: str
    client_codes_csv: str
    email_pdf: str
    attachment_pdf: str
    catalogued_zip: str
    unclassified_zip: str
    manifest_json: str


@dataclass
class AttachmentBundleResult:
    source: str
    output_dir: str
    attachment_count: int
    catalogued_count: int
    unclassified_count: int
    excluded_media_count: int
    duplicate_count: int
    attachment_csv: str
    client_codes_csv: str
    attachment_pdf: str
    catalogued_zip: str
    unclassified_zip: str
    manifest_json: str


def _decode_header(value: str | None) -> str:
    if not value:
        return ""
    try:
        return str(make_header(decode_header(value)))
    except Exception:
        return str(value)


def _decode_bytes(raw: bytes, charset: str | None) -> str:
    candidates = [charset, "utf-8", "cp1252", "latin-1"]
    for candidate in candidates:
        if not candidate:
            continue
        try:
            return raw.decode(candidate, errors="replace")
        except LookupError:
            continue
    return raw.decode("utf-8", errors="replace")


def _strip_html(raw_html: str) -> str:
    parser = _HTMLTextExtractor()
    try:
        parser.feed(raw_html)
        return html.unescape(parser.text())
    except Exception:
        return re.sub(r"<[^>]+>", " ", raw_html)


def message_body(message: Message, max_chars: int = 120_000) -> str:
    plain: list[str] = []
    html_parts: list[str] = []
    parts: Iterable[Message] = message.walk() if message.is_multipart() else [message]
    for part in parts:
        disposition = (part.get_content_disposition() or "").lower()
        if disposition == "attachment" or part.get_filename():
            continue
        content_type = part.get_content_type().lower()
        if content_type not in {"text/plain", "text/html"}:
            continue
        payload = part.get_payload(decode=True)
        if payload is None:
            payload_obj = part.get_payload()
            text = payload_obj if isinstance(payload_obj, str) else ""
        else:
            text = _decode_bytes(payload, part.get_content_charset())
        if content_type == "text/plain":
            plain.append(text)
        else:
            html_parts.append(_strip_html(text))
        if sum(len(x) for x in plain) >= max_chars:
            break
    body = "\n".join(plain) if plain else "\n".join(html_parts)
    body = re.sub(r"[ \t]+", " ", body)
    body = re.sub(r"\n{3,}", "\n\n", body)
    return body[:max_chars]


def clean_summary(text: str, limit: int = 650) -> str:
    compact = re.sub(r"\s+", " ", text or "").strip()
    return compact[:limit]


def normalize_company(value: str) -> str:
    s = (value or "").upper().replace("&", " E ")
    replacements = {
        "SOCIETA' A RESPONSABILITA' LIMITATA": " SRL ",
        "SOCIETÀ A RESPONSABILITÀ LIMITATA": " SRL ",
        "SOCIETA A RESPONSABILITA LIMITATA": " SRL ",
        "SOCIETA' PER AZIONI": " SPA ",
        "SOCIETÀ PER AZIONI": " SPA ",
        "SOCIETA PER AZIONI": " SPA ",
        "S.R.L.S.": " SRLS ",
        "S.R.L.": " SRL ",
        "S R L": " SRL ",
        "S.P.A.": " SPA ",
        "S P A": " SPA ",
        "S.N.C.": " SNC ",
        "S.A.S.": " SAS ",
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    s = re.sub(r"\bUNIPERSONALE\b", " ", s)
    return re.sub(r"[^A-Z0-9]+", "", s)


def display_company(value: str) -> str:
    s = re.sub(r"\s+", " ", (value or "").strip(" .,-_"))
    s = re.sub(r"S\.?\s*P\.?\s*A\.?$", "SPA", s, flags=re.IGNORECASE)
    s = re.sub(r"S\.?\s*R\.?\s*L\.?\s*S\.?$", "SRLS", s, flags=re.IGNORECASE)
    s = re.sub(r"S\.?\s*R\.?\s*L\.?$", "SRL", s, flags=re.IGNORECASE)
    return s.upper() if s else "DA_ARCHIVIARE"


def safe_component(value: str, fallback: str = "file") -> str:
    cleaned = SAFE_COMPONENT_RE.sub("_", value or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    if not cleaned:
        cleaned = fallback
    return cleaned[:150]


def _known_client_match(corpus: str, known_clients: Sequence[str]) -> tuple[str | None, float]:
    norm_corpus = normalize_company(corpus)
    if not norm_corpus:
        return None, 0.0
    best_name: str | None = None
    best_score = 0.0
    for client in known_clients:
        norm_client = normalize_company(client)
        if len(norm_client) < 3:
            continue
        if norm_client in norm_corpus:
            score = min(1.0, 0.90 + min(len(norm_client), 20) / 200)
        else:
            positions = [0]
            if len(norm_corpus) > len(norm_client):
                positions.extend(range(0, min(len(norm_corpus), 4000), max(len(norm_client), 20)))
            score = 0.0
            for pos in positions[:120]:
                window = norm_corpus[pos : pos + max(len(norm_client) + 8, 20)]
                score = max(score, SequenceMatcher(None, norm_client, window).ratio())
        if score > best_score:
            best_name, best_score = client, score
    if best_score >= 0.84:
        return best_name, best_score
    return None, best_score


def identify_client(corpus: str, known_clients: Sequence[str] | None = None) -> tuple[str, float, str]:
    known_clients = [str(x).strip() for x in (known_clients or []) if str(x).strip()]
    matched, score = _known_client_match(corpus[:60_000], known_clients)
    if matched:
        return display_company(matched), score, "anagrafica"
    candidates = []
    for match in COMPANY_SUFFIX_RE.finditer(corpus[:30_000]):
        value = display_company(match.group(1))
        if value not in candidates:
            candidates.append(value)
    if candidates:
        candidates.sort(key=lambda x: (len(x) > 70, len(x)))
        return candidates[0], 0.68, "testo"
    return "DA_ARCHIVIARE", 0.0, "nessun_riferimento"


def _load_client_registry(path: Path, known_clients: Sequence[str] | None = None) -> dict[str, dict[str, object]]:
    registry: dict[str, dict[str, object]] = {}
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                registry = data
        except Exception:
            registry = {}
    next_code = max([int(v.get("code", 0) or 0) for v in registry.values()] or [0]) + 1
    for client in known_clients or []:
        display = display_company(str(client))
        norm = normalize_company(display)
        if not norm or norm in registry:
            continue
        registry[norm] = {"client": display, "code": next_code}
        next_code += 1
    return registry


def _client_folder(client: str, registry: dict[str, dict[str, object]]) -> tuple[str, str]:
    if client == "DA_ARCHIVIARE":
        return "DA_ARCHIVIARE", "000"
    norm = normalize_company(client)
    if norm not in registry:
        next_code = max([int(v.get("code", 0) or 0) for v in registry.values()] or [0]) + 1
        registry[norm] = {"client": display_company(client), "code": next_code}
    item = registry[norm]
    code = f"{int(item.get('code', 0) or 0):03d}"
    return f"{safe_component(str(item.get('client') or client), 'CLIENTE')} {code}", code


def _write_registry(registry: dict[str, dict[str, object]], json_path: Path, csv_path: Path) -> None:
    json_path.write_text(json.dumps(registry, ensure_ascii=False, indent=2), encoding="utf-8")
    rows = sorted(registry.values(), key=lambda x: int(x.get("code", 0) or 0))
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=["Codice", "Cliente", "Cartella"])
        writer.writeheader()
        for item in rows:
            code = f"{int(item.get('code', 0) or 0):03d}"
            client = str(item.get("client") or "")
            writer.writerow({"Codice": code, "Cliente": client, "Cartella": f"{client} {code}"})


def _message_datetime(message: Message) -> str:
    raw = message.get("Date")
    if not raw:
        return ""
    try:
        dt = parsedate_to_datetime(raw)
        if dt is None:
            return _decode_header(raw)
        return dt.astimezone().isoformat(timespec="seconds") if dt.tzinfo else dt.isoformat(timespec="seconds")
    except Exception:
        return _decode_header(raw)


def _addresses(message: Message, header: str) -> str:
    values = message.get_all(header, [])
    parsed = getaddresses(values)
    rendered = []
    for name, address in parsed:
        display = _decode_header(name)
        rendered.append(f"{display} <{address}>" if display and address else address or display)
    return "; ".join(x for x in rendered if x)


def _iter_attachments(message: Message):
    parts: Iterable[Message] = message.walk() if message.is_multipart() else [message]
    for part in parts:
        filename = part.get_filename()
        disposition = (part.get_content_disposition() or "").lower()
        if not filename and disposition != "attachment":
            continue
        payload = part.get_payload(decode=True)
        if payload is None:
            continue
        filename = _decode_header(filename) if filename else "allegato.bin"
        mime_type = part.get_content_type() or mimetypes.guess_type(filename)[0] or "application/octet-stream"
        yield filename, mime_type, payload


def _unique_destination(directory: Path, filename: str, sha256: str) -> Path:
    filename = safe_component(filename, "allegato.bin")
    candidate = directory / filename
    if not candidate.exists():
        return candidate
    stem, suffix = candidate.stem, candidate.suffix
    return directory / f"{stem}_{sha256[:10]}{suffix}"


def _is_media(filename: str, mime_type: str = "") -> bool:
    ext = Path(filename).suffix.lower()
    return ext in MEDIA_EXTENSIONS or mime_type.startswith("image/") or mime_type.startswith("video/")


def _zip_directory(source_dir: Path, zip_path: Path) -> str:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, allowZip64=True) as archive:
        if source_dir.exists():
            for path in source_dir.rglob("*"):
                if path.is_file():
                    archive.write(path, path.relative_to(source_dir))
    return str(zip_path)


def _pdf_text(value: object, limit: int = 240) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()[:limit]
    return html.escape(text)


def _build_catalog_pdf(rows: list[dict[str, object]], path: Path, title: str, columns: list[tuple[str, str, float]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CatalogTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=15, leading=18, alignment=TA_CENTER, spaceAfter=8)
    body_style = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=6.7, leading=8.2)
    header_style = ParagraphStyle("Header", parent=body_style, fontName="Helvetica-Bold", textColor=colors.white)
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), leftMargin=8 * mm, rightMargin=8 * mm, topMargin=8 * mm, bottomMargin=8 * mm)
    story = [Paragraph(title, title_style), Paragraph(f"Generato: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Righe: {len(rows)}", body_style), Spacer(1, 4 * mm)]
    if not rows:
        story.append(Paragraph("Nessun dato disponibile.", body_style))
        doc.build(story)
        return
    widths = [width * mm for _, _, width in columns]
    data = [[Paragraph(label, header_style) for _, label, _ in columns]]
    for row in rows:
        data.append([Paragraph(_pdf_text(row.get(key, "")), body_style) for key, _, _ in columns])
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#234A7B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C2CC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FA")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
        ("TOPPADDING", (0, 0), (-1, -1), 2.2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2),
    ]))
    story.append(table)
    doc.build(story)


def _write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def import_mbox(mbox_path: str | os.PathLike[str], output_dir: str | os.PathLike[str], known_clients: Sequence[str] | None = None, *, exclude_media_from_physical_archive: bool = True, create_zip_archives: bool = True, max_messages: int | None = None, progress: Callable[[int, int], None] | None = None) -> ImportResult:
    source = Path(mbox_path).expanduser().resolve()
    if not source.exists() or not source.is_file():
        raise FileNotFoundError(f"Archivio MBOX non trovato: {source}")
    root = Path(output_dir).expanduser().resolve()
    root.mkdir(parents=True, exist_ok=True)
    catalogued_root = root / "ALLEGATI_CATALOGATI"
    unclassified_root = root / "ALLEGATI_DA_ARCHIVIARE"
    catalogued_root.mkdir(parents=True, exist_ok=True)
    unclassified_root.mkdir(parents=True, exist_ok=True)
    registry_json = root / "clienti_codici.json"
    registry_csv = root / "clienti_codici.csv"
    registry = _load_client_registry(registry_json, known_clients)
    email_rows: list[dict[str, object]] = []
    attachment_rows: list[dict[str, object]] = []
    seen_sha: dict[str, str] = {}
    counters = Counter()
    box = mailbox.mbox(str(source), create=False)
    for index, message in enumerate(box, start=1):
        if max_messages is not None and index > max_messages:
            break
        subject = _decode_header(message.get("Subject"))
        sender = _addresses(message, "From")
        recipients = _addresses(message, "To")
        cc = _addresses(message, "Cc")
        body = message_body(message)
        date_value = _message_datetime(message)
        context = "\n".join([subject, sender, recipients, cc, body[:40_000]])
        client, confidence, client_source = identify_client(context, known_clients)
        folder_name, client_code = _client_folder(client, registry)
        attachment_names: list[str] = []
        message_attachment_count = 0
        for attachment_index, (filename, mime_type, raw) in enumerate(_iter_attachments(message), start=1):
            counters["attachments"] += 1
            message_attachment_count += 1
            attachment_names.append(filename)
            sha = hashlib.sha256(raw).hexdigest()
            media = _is_media(filename, mime_type)
            duplicate_of = seen_sha.get(sha, "")
            duplicate = bool(duplicate_of)
            if duplicate:
                counters["duplicates"] += 1
            else:
                seen_sha[sha] = f"msg:{index}/att:{attachment_index}"
            attachment_client = client
            attachment_confidence = confidence
            attachment_source = client_source
            if attachment_client == "DA_ARCHIVIARE":
                attachment_client, attachment_confidence, attachment_source = identify_client(f"{filename}\n{subject}\n{body[:12_000]}", known_clients)
            attachment_folder, attachment_code = _client_folder(attachment_client, registry)
            excluded = bool(media and exclude_media_from_physical_archive)
            physical_path = ""
            if excluded:
                counters["excluded_media"] += 1
            elif duplicate:
                physical_path = f"DUPLICATO:{duplicate_of}"
            else:
                target_base = catalogued_root if attachment_client != "DA_ARCHIVIARE" else unclassified_root
                target_dir = target_base / attachment_folder if attachment_client != "DA_ARCHIVIARE" else target_base
                target_dir.mkdir(parents=True, exist_ok=True)
                target = _unique_destination(target_dir, filename, sha)
                target.write_bytes(raw)
                physical_path = str(target.relative_to(root))
                if attachment_client == "DA_ARCHIVIARE":
                    counters["unclassified"] += 1
                else:
                    counters["catalogued"] += 1
            attachment_rows.append({"Email #": index, "Data e ora": date_value, "Cliente": attachment_client, "Codice cliente": attachment_code, "Cartella cliente": attachment_folder, "Nome file originale": filename, "MIME": mime_type, "Dimensione byte": len(raw), "SHA-256": sha, "Duplicato": "SI" if duplicate else "NO", "Escluso foto/video": "SI" if excluded else "NO", "Percorso fisico": physical_path, "Oggetto email": subject, "Mittente": sender, "Confidenza cliente": round(float(attachment_confidence), 3), "Metodo associazione": attachment_source})
        email_rows.append({"Email #": index, "Data e ora": date_value, "Message-ID": _decode_header(message.get("Message-ID")), "Mittente": sender, "Destinatari": recipients, "CC": cc, "Oggetto": subject, "Sintesi": clean_summary(body), "Cliente": client, "Codice cliente": client_code, "Cartella cliente": folder_name, "Confidenza cliente": round(float(confidence), 3), "Metodo associazione": client_source, "N. allegati": message_attachment_count, "Allegati": " | ".join(attachment_names)})
        counters["emails"] += 1
        if progress:
            progress(index, counters["attachments"])
    _write_registry(registry, registry_json, registry_csv)
    email_csv = root / "catalogo_email.csv"
    attachment_csv = root / "catalogo_allegati.csv"
    _write_csv(email_csv, email_rows, ["Email #", "Data e ora", "Message-ID", "Mittente", "Destinatari", "CC", "Oggetto", "Sintesi", "Cliente", "Codice cliente", "Cartella cliente", "Confidenza cliente", "Metodo associazione", "N. allegati", "Allegati"])
    _write_csv(attachment_csv, attachment_rows, ["Email #", "Data e ora", "Cliente", "Codice cliente", "Cartella cliente", "Nome file originale", "MIME", "Dimensione byte", "SHA-256", "Duplicato", "Escluso foto/video", "Percorso fisico", "Oggetto email", "Mittente", "Confidenza cliente", "Metodo associazione"])
    email_pdf = root / f"Catalogo_Email_{safe_component(source.stem, 'MBOX')}.pdf"
    attachment_pdf = root / f"Catalogo_Allegati_{safe_component(source.stem, 'MBOX')}.pdf"
    _build_catalog_pdf(email_rows, email_pdf, f"Catalogo email MBOX - {html.escape(source.name)}", [("Email #", "#", 8), ("Data e ora", "Data/ora", 31), ("Cliente", "Cliente", 36), ("Mittente", "Mittente", 43), ("Oggetto", "Oggetto", 62), ("Sintesi", "Sintesi", 90), ("Allegati", "Allegati", 55)])
    _build_catalog_pdf(attachment_rows, attachment_pdf, f"Catalogo allegati MBOX - {html.escape(source.name)}", [("Email #", "Email", 10), ("Data e ora", "Data/ora", 31), ("Cliente", "Cliente", 40), ("Codice cliente", "Cod.", 12), ("Nome file originale", "File originale", 65), ("Oggetto email", "Oggetto email", 70), ("Percorso fisico", "Percorso", 70)])
    catalogued_zip = root / "Allegati_Catalogati.zip"
    unclassified_zip = root / "Allegati_DA_ARCHIVIARE.zip"
    if create_zip_archives:
        _zip_directory(catalogued_root, catalogued_zip)
        _zip_directory(unclassified_root, unclassified_zip)
    manifest = {"source": str(source), "created_at": datetime.now().isoformat(timespec="seconds"), "email_count": counters["emails"], "attachment_count": counters["attachments"], "catalogued_count": counters["catalogued"], "unclassified_count": counters["unclassified"], "excluded_media_count": counters["excluded_media"], "duplicate_count": counters["duplicates"], "exclude_media_from_physical_archive": exclude_media_from_physical_archive, "outputs": {"email_csv": str(email_csv), "attachment_csv": str(attachment_csv), "client_codes_csv": str(registry_csv), "email_pdf": str(email_pdf), "attachment_pdf": str(attachment_pdf), "catalogued_zip": str(catalogued_zip) if create_zip_archives else "", "unclassified_zip": str(unclassified_zip) if create_zip_archives else ""}}
    manifest_json = root / "manifest.json"
    manifest_json.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return ImportResult(source=str(source), output_dir=str(root), email_count=counters["emails"], attachment_count=counters["attachments"], catalogued_count=counters["catalogued"], unclassified_count=counters["unclassified"], excluded_media_count=counters["excluded_media"], duplicate_count=counters["duplicates"], email_csv=str(email_csv), attachment_csv=str(attachment_csv), client_codes_csv=str(registry_csv), email_pdf=str(email_pdf), attachment_pdf=str(attachment_pdf), catalogued_zip=str(catalogued_zip) if create_zip_archives else "", unclassified_zip=str(unclassified_zip) if create_zip_archives else "", manifest_json=str(manifest_json))


def _document_hint(path: Path, max_chars: int = 18_000) -> str:
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            return "\n".join((page.extract_text() or "") for page in reader.pages[:3])[:max_chars]
        if ext == ".docx":
            from docx import Document
            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs[:120])[:max_chars]
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            pieces: list[str] = []
            for ws in wb.worksheets[:4]:
                for row in ws.iter_rows(max_row=40, values_only=True):
                    pieces.append(" ".join(str(v) for v in row if v is not None))
                    if sum(len(x) for x in pieces) > max_chars:
                        break
            wb.close()
            return "\n".join(pieces)[:max_chars]
        if ext in {".txt", ".csv", ".md", ".xml", ".json"}:
            return path.read_text(encoding="utf-8", errors="replace")[:max_chars]
    except Exception:
        return ""
    return ""


def _safe_extract_zip(zip_path: Path, destination: Path) -> None:
    destination.mkdir(parents=True, exist_ok=True)
    base = destination.resolve()
    with zipfile.ZipFile(zip_path) as archive:
        for member in archive.infolist():
            candidate = (destination / member.filename).resolve()
            try:
                candidate.relative_to(base)
            except ValueError:
                continue
            if member.is_dir():
                candidate.mkdir(parents=True, exist_ok=True)
                continue
            candidate.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(member) as src, candidate.open("wb") as dst:
                shutil.copyfileobj(src, dst, length=1024 * 1024)


def classify_attachment_bundle(source_path: str | os.PathLike[str], output_dir: str | os.PathLike[str], known_clients: Sequence[str] | None = None, *, exclude_media_from_physical_archive: bool = True, create_zip_archives: bool = True, progress: Callable[[int, int], None] | None = None) -> AttachmentBundleResult:
    source = Path(source_path).expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(f"Sorgente allegati non trovata: {source}")
    root = Path(output_dir).expanduser().resolve()
    root.mkdir(parents=True, exist_ok=True)
    catalogued_root = root / "ALLEGATI_CATALOGATI"
    unclassified_root = root / "ALLEGATI_DA_ARCHIVIARE"
    catalogued_root.mkdir(parents=True, exist_ok=True)
    unclassified_root.mkdir(parents=True, exist_ok=True)
    temp_dir: tempfile.TemporaryDirectory[str] | None = None
    if source.is_file() and source.suffix.lower() == ".zip":
        temp_dir = tempfile.TemporaryDirectory(prefix="financeplus_attachments_")
        scan_root = Path(temp_dir.name)
        _safe_extract_zip(source, scan_root)
    elif source.is_dir():
        scan_root = source
    else:
        raise ValueError("La sorgente deve essere una cartella oppure un file ZIP.")
    registry_json = root / "clienti_codici.json"
    registry_csv = root / "clienti_codici.csv"
    registry = _load_client_registry(registry_json, known_clients)
    rows: list[dict[str, object]] = []
    seen_sha: dict[str, str] = {}
    counters = Counter()
    files = [p for p in scan_root.rglob("*") if p.is_file() and not p.name.startswith(".")]
    for index, path in enumerate(files, start=1):
        counters["attachments"] += 1
        mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        media = _is_media(path.name, mime_type)
        raw_sha = hashlib.sha256()
        with path.open("rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                raw_sha.update(chunk)
        sha = raw_sha.hexdigest()
        duplicate_of = seen_sha.get(sha, "")
        duplicate = bool(duplicate_of)
        if duplicate:
            counters["duplicates"] += 1
        else:
            seen_sha[sha] = str(path.relative_to(scan_root))
        extracted = _document_hint(path)
        hint = "\n".join([str(path.relative_to(scan_root)), path.name, extracted])
        client, confidence, method = identify_client(hint, known_clients)
        folder_name, code = _client_folder(client, registry)
        excluded = bool(media and exclude_media_from_physical_archive)
        physical_path = ""
        if excluded:
            counters["excluded_media"] += 1
        elif duplicate:
            physical_path = f"DUPLICATO:{duplicate_of}"
        else:
            target_base = catalogued_root if client != "DA_ARCHIVIARE" else unclassified_root
            target_dir = target_base / folder_name if client != "DA_ARCHIVIARE" else target_base
            target_dir.mkdir(parents=True, exist_ok=True)
            target = _unique_destination(target_dir, path.name, sha)
            shutil.copy2(path, target)
            physical_path = str(target.relative_to(root))
            if client == "DA_ARCHIVIARE":
                counters["unclassified"] += 1
            else:
                counters["catalogued"] += 1
        rows.append({"#": index, "Cliente": client, "Codice cliente": code, "Cartella cliente": folder_name, "Nome file originale": path.name, "Percorso origine": str(path.relative_to(scan_root)), "MIME": mime_type, "Dimensione byte": path.stat().st_size, "SHA-256": sha, "Duplicato": "SI" if duplicate else "NO", "Escluso foto/video": "SI" if excluded else "NO", "Percorso fisico": physical_path, "Confidenza cliente": round(float(confidence), 3), "Metodo associazione": method, "Estratto contenuto": clean_summary(extracted, 450)})
        if progress:
            progress(index, len(files))
    _write_registry(registry, registry_json, registry_csv)
    attachment_csv = root / "catalogo_allegati_esportati.csv"
    _write_csv(attachment_csv, rows, ["#", "Cliente", "Codice cliente", "Cartella cliente", "Nome file originale", "Percorso origine", "MIME", "Dimensione byte", "SHA-256", "Duplicato", "Escluso foto/video", "Percorso fisico", "Confidenza cliente", "Metodo associazione", "Estratto contenuto"])
    attachment_pdf = root / "Catalogo_Allegati_Esportati.pdf"
    _build_catalog_pdf(rows, attachment_pdf, "Catalogo allegati esportati da MBOX Viewer", [("#", "#", 9), ("Cliente", "Cliente", 42), ("Codice cliente", "Cod.", 13), ("Nome file originale", "File originale", 62), ("Estratto contenuto", "Sintesi contenuto", 90), ("Percorso fisico", "Percorso catalogato", 72)])
    catalogued_zip = root / "Allegati_Catalogati.zip"
    unclassified_zip = root / "Allegati_DA_ARCHIVIARE.zip"
    if create_zip_archives:
        _zip_directory(catalogued_root, catalogued_zip)
        _zip_directory(unclassified_root, unclassified_zip)
    manifest = {"source": str(source), "created_at": datetime.now().isoformat(timespec="seconds"), "attachment_count": counters["attachments"], "catalogued_count": counters["catalogued"], "unclassified_count": counters["unclassified"], "excluded_media_count": counters["excluded_media"], "duplicate_count": counters["duplicates"]}
    manifest_json = root / "manifest_allegati.json"
    manifest_json.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    if temp_dir is not None:
        temp_dir.cleanup()
    return AttachmentBundleResult(source=str(source), output_dir=str(root), attachment_count=counters["attachments"], catalogued_count=counters["catalogued"], unclassified_count=counters["unclassified"], excluded_media_count=counters["excluded_media"], duplicate_count=counters["duplicates"], attachment_csv=str(attachment_csv), client_codes_csv=str(registry_csv), attachment_pdf=str(attachment_pdf), catalogued_zip=str(catalogued_zip) if create_zip_archives else "", unclassified_zip=str(unclassified_zip) if create_zip_archives else "", manifest_json=str(manifest_json))


def result_as_dict(result: ImportResult | AttachmentBundleResult) -> dict[str, object]:
    return asdict(result)
